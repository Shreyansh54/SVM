import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from auth import require_role
from fpdf import FPDF
from openpyxl import Workbook
import models

router = APIRouter(prefix="/api/export", tags=["Export"])


def clean_pdf_text(text):
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    # Replace common unicode characters that cause issues in core PDF fonts
    text = text.replace("—", "-").replace("–", "-").replace("₹", "Rs. ").replace("’", "'").replace("“", '"').replace("”", '"')
    # Encode as latin-1, replacing unsupported characters with '?' or similar fallback
    return text.encode("latin-1", "replace").decode("latin-1")


class VolloraReport(FPDF):
    def header(self):
        # Draw background decoration or branding header
        self.set_y(10)
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(10, 55, 58) # Deep Forest Teal
        self.cell(0, 8, "SHREYANSH VOLLORA PVT LTD", align="C", new_x="LMARGIN", new_y="NEXT")
        
        self.set_font("Helvetica", "B", 8)
        self.set_text_color(74, 109, 113) # Muted Teal-Gray
        self.cell(0, 4, "Every Step GUIDED BY CARE", align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(2)
        
        # Dual color accent bar
        # Forest teal line
        self.set_fill_color(10, 55, 58)
        self.rect(10, 26, 190, 1.5, "F")
        # Turquoise line
        self.set_fill_color(20, 168, 156)
        self.rect(10, 27.5, 190, 0.5, "F")
        
        # Set the Y position for page content so it starts cleanly below the header bar!
        self.set_y(32)

    def footer(self):
        self.set_y(-15)
        # Line above footer
        self.set_draw_color(213, 229, 228)
        self.line(10, 282, 200, 282)
        
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(74, 109, 113)
        # Left side text
        self.cell(95, 10, "SHREYANSH VOLLORA PVT LTD  Confidential", align="L")
        # Right side text
        self.cell(95, 10, f"Page {self.page_no()}/{{nb}}", align="R")

    def section_title(self, title):
        self.ln(4)
        self.set_font("Helvetica", "B", 11)
        self.set_fill_color(10, 55, 58) # Deep Forest Teal
        self.set_text_color(255, 255, 255)
        # Write padded cell
        self.cell(0, 8, clean_pdf_text(f"   {title.upper()}"), fill=True, new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0, 0, 0)
        self.ln(2)

    def table_header(self, cols, widths):
        self.set_font("Helvetica", "B", 9)
        self.set_fill_color(227, 239, 239)
        self.set_text_color(10, 55, 58)
        self.set_draw_color(213, 229, 228)
        for col, w in zip(cols, widths):
            self.cell(w, 8, clean_pdf_text(col), border=1, fill=True, align="C")
        self.ln()

    def table_row(self, values, widths, row_idx=0):
        self.set_font("Helvetica", "", 8.5)
        if row_idx % 2 == 1:
            self.set_fill_color(245, 249, 249)
            fill = True
        else:
            fill = False
        self.set_text_color(26, 61, 64)
        self.set_draw_color(225, 236, 235)
        for val, w in zip(values, widths):
            # Safe truncation and string conversion
            val_str = str(val)[:30] if val is not None else ""
            self.cell(w, 7.5, clean_pdf_text(val_str), border=1, fill=fill, align="C")
        self.ln()


@router.get("/pdf/all")
def export_all_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager", "hr"))
):
    pdf = VolloraReport()
    pdf.set_margins(10, 32, 10)
    pdf.set_auto_page_break(True, margin=15)
    pdf.alias_nb_pages()

    # Employees
    employees = db.query(models.Employee).all()
    pdf.add_page()
    pdf.section_title("Employees")
    cols = ["ID", "Name", "Email", "Phone", "Salary", "Role", "Active"]
    widths = [15, 35, 45, 30, 25, 20, 20] # Sum: 190
    pdf.table_header(cols, widths)
    for idx, e in enumerate(employees):
        pdf.table_row([e.id, e.name, e.email or "", e.phone or "", f"{e.salary_per_month:.0f}", e.role, "Yes" if e.is_active else "No"], widths, idx)

    # Stockists
    stockists = db.query(models.Stockist).all()
    pdf.add_page()
    pdf.section_title("Stockists")
    cols = ["ID", "Name", "Location", "Contact Person", "Phone"]
    widths = [15, 40, 45, 45, 45] # Sum: 190
    pdf.table_header(cols, widths)
    for idx, s in enumerate(stockists):
        pdf.table_row([s.id, s.name, s.location or "", s.contact_person or "", s.phone or ""], widths, idx)

    # Medicines
    products = db.query(models.Product).all()
    pdf.add_page()
    pdf.section_title("Medicine Database")
    cols = ["Name", "Generic Name", "Category", "Mfr", "Sch", "Price"]
    widths = [35, 40, 30, 35, 20, 30] # Sum: 190
    pdf.table_header(cols, widths)
    for idx, p in enumerate(products):
        pdf.table_row([p.name, p.generic_name or "", p.category or "", p.manufacturer or "", p.schedule_type or "", f"{p.price:.2f}"], widths, idx)

    # Stock
    stocks = db.query(models.Stock).all()
    pdf.add_page()
    pdf.section_title("Stock Inventory")
    cols = ["ID", "Stockist", "Product", "Batch", "Qty", "Updated"]
    widths = [15, 40, 40, 35, 20, 40] # Sum: 190
    pdf.table_header(cols, widths)
    for idx, s in enumerate(stocks):
        pdf.table_row([
            s.id,
            s.stockist.name if s.stockist else "",
            s.product.name if s.product else "",
            s.batch.batch_number if s.batch else "—",
            s.quantity,
            s.last_updated.strftime("%Y-%m-%d") if s.last_updated else ""
        ], widths, idx)

    # Sales
    sales = db.query(models.Sale).all()
    pdf.add_page()
    pdf.section_title("Sales")
    cols = ["ID", "Employee", "Channel", "Product", "Qty", "Disc%", "Amount"]
    widths = [15, 30, 35, 35, 20, 20, 35] # Sum: 190
    pdf.table_header(cols, widths)
    for idx, s in enumerate(sales):
        if s.sale_type == "doctor":
            channel = f"Dr. {s.doctor.name}" if s.doctor else "Doctor"
        else:
            channel = s.stockist.name if s.stockist else "Stockist"
        pdf.table_row([
            s.id,
            s.employee.name if s.employee else "",
            channel,
            s.product.name if s.product else "",
            s.quantity_sold,
            f"{s.discount_percentage:.0f}%" if s.discount_percentage else "—",
            f"{s.total_amount:.2f}",
        ], widths, idx)

    # Batches
    batches = db.query(models.Batch).all()
    if batches:
        pdf.add_page()
        pdf.section_title("Batches")
        cols = ["Batch#", "Product", "Mfg", "Expiry", "MRP", "GST%", "Status"]
        widths = [30, 35, 25, 25, 25, 25, 25] # Sum: 190
        pdf.table_header(cols, widths)
        for idx, b in enumerate(batches):
            pdf.table_row([
                b.batch_number,
                b.product.name if b.product else "",
                b.manufacturing_date.strftime("%Y-%m-%d") if b.manufacturing_date else "",
                b.expiry_date.strftime("%Y-%m-%d") if b.expiry_date else "",
                f"{b.mrp:.2f}",
                f"{b.gst_percentage:.1f}",
                b.status.upper()
            ], widths, idx)

    # Doctors
    doctors = db.query(models.Doctor).all()
    if doctors:
        pdf.add_page()
        pdf.section_title("Doctors")
        cols = ["ID", "Name", "Specialization", "Hospital", "Phone", "Location"]
        widths = [15, 35, 35, 35, 35, 35] # Sum: 190
        pdf.table_header(cols, widths)
        for idx, d in enumerate(doctors):
            pdf.table_row([d.id, d.name, d.specialization or "", d.hospital or "", d.phone or "", d.location or ""], widths, idx)

    # Doctor Orders
    orders = db.query(models.DoctorOrder).all()
    if orders:
        pdf.add_page()
        pdf.section_title("Doctor Orders")
        cols = ["ID", "Doctor", "Employee", "Product", "Qty", "Date"]
        widths = [15, 35, 35, 35, 30, 40] # Sum: 190
        pdf.table_header(cols, widths)
        for idx, o in enumerate(orders):
            pdf.table_row([
                o.id,
                o.doctor.name if o.doctor else "",
                o.employee.name if o.employee else "",
                o.product.name if o.product else "",
                o.quantity,
                o.date.strftime("%Y-%m-%d") if o.date else ""
            ], widths, idx)

    buf = io.BytesIO(pdf.output())
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=SHREYANSH_VOLLORA_Report.pdf"})


@router.get("/excel/all")
def export_all_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager", "hr"))
):
    wb = Workbook()

    # Employees sheet
    ws = wb.active
    ws.title = "Employees"
    ws.append(["ID", "Name", "Email", "Phone", "Salary", "Joining Date", "Role", "Active"])
    for e in db.query(models.Employee).all():
        ws.append([e.id, e.name, e.email, e.phone, e.salary_per_month, str(e.joining_date), e.role, e.is_active])

    # Stockists sheet
    ws2 = wb.create_sheet("Stockists")
    ws2.append(["ID", "Name", "Location", "Contact Person", "Phone"])
    for s in db.query(models.Stockist).all():
        ws2.append([s.id, s.name, s.location, s.contact_person, s.phone])

    # Medicines sheet
    ws3 = wb.create_sheet("Medicines")
    ws3.append(["ID", "Name", "Generic Name", "Composition", "Dosage", "Category", "Packaging", "Manufacturer", "Schedule", "Price"])
    for p in db.query(models.Product).all():
        ws3.append([p.id, p.name, p.generic_name, p.composition, p.dosage, p.category, p.packaging, p.manufacturer, p.schedule_type, p.price])

    # Stock sheet
    ws4 = wb.create_sheet("Stock")
    ws4.append(["ID", "Stockist", "Product", "Batch", "Quantity", "Last Updated"])
    for s in db.query(models.Stock).all():
        ws4.append([s.id, s.stockist.name if s.stockist else "", s.product.name if s.product else "", s.batch.batch_number if s.batch else "", s.quantity, str(s.last_updated) if s.last_updated else ""])

    # Batches sheet
    wsB = wb.create_sheet("Batches")
    wsB.append(["ID", "Product", "Batch Number", "Mfg Date", "Expiry Date", "MRP", "GST%", "Purchase Price", "Status", "Notes"])
    for b in db.query(models.Batch).all():
        wsB.append([b.id, b.product.name if b.product else "", b.batch_number, str(b.manufacturing_date), str(b.expiry_date), b.mrp, b.gst_percentage, b.purchase_price, b.status, b.notes])

    # Sales sheet
    ws5 = wb.create_sheet("Sales")
    ws5.append(["ID", "Employee", "Channel", "Stockist/Doctor", "Product", "Batch", "Qty Sold", "Discount%", "Amount", "Date"])
    for s in db.query(models.Sale).all():
        channel_name = ""
        if s.sale_type == "doctor":
            channel_name = f"Dr. {s.doctor.name}" if s.doctor else ""
        else:
            channel_name = s.stockist.name if s.stockist else ""
        ws5.append([s.id, s.employee.name if s.employee else "", s.sale_type or "stockist", channel_name, s.product.name if s.product else "", s.batch.batch_number if s.batch else "", s.quantity_sold, s.discount_percentage or 0, s.total_amount, str(s.date) if s.date else ""])

    # Doctors sheet
    ws6 = wb.create_sheet("Doctors")
    ws6.append(["ID", "Name", "Specialization", "Hospital", "Phone", "Location"])
    for d in db.query(models.Doctor).all():
        ws6.append([d.id, d.name, d.specialization, d.hospital, d.phone, d.location])

    # Doctor Orders sheet
    ws7 = wb.create_sheet("Doctor Orders")
    ws7.append(["ID", "Doctor", "Employee", "Product", "Quantity", "Date", "Notes"])
    for o in db.query(models.DoctorOrder).all():
        ws7.append([o.id, o.doctor.name if o.doctor else "", o.employee.name if o.employee else "", o.product.name if o.product else "", o.quantity, str(o.date) if o.date else "", o.notes])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=SHREYANSH_VOLLORA_Report.xlsx"})
