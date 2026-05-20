from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from database import get_db
from auth import require_role
from openpyxl import load_workbook
from datetime import datetime
import models
import io

router = APIRouter(prefix="/api/upload", tags=["Upload"])


@router.post("/employees")
def upload_employees(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager"))
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx) are supported")

    contents = file.file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active

    added = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            name = str(row[0]).strip()
            email = str(row[1]).strip() if row[1] else f"{name.lower().replace(' ', '.')}@company.com"
            phone = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            salary = float(row[3]) if len(row) > 3 and row[3] else 0
            joining_date_raw = row[4] if len(row) > 4 and row[4] else datetime.now().date()
            role = str(row[5]).strip().lower() if len(row) > 5 and row[5] else "employee"

            if isinstance(joining_date_raw, str):
                joining_date = datetime.strptime(joining_date_raw, "%Y-%m-%d").date()
            elif isinstance(joining_date_raw, datetime):
                joining_date = joining_date_raw.date()
            else:
                joining_date = joining_date_raw

            # Skip if email already exists
            existing = db.query(models.Employee).filter(models.Employee.email == email).first()
            if existing:
                errors.append(f"Row {i}: Email {email} already exists, skipped")
                continue

            emp = models.Employee(
                name=name, email=email, phone=phone,
                salary_per_month=salary, joining_date=joining_date,
                role=role, is_active=True
            )
            db.add(emp)
            added += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()
    return {"message": f"Successfully added {added} employees", "errors": errors}


@router.post("/stockists")
def upload_stockists(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager"))
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx) are supported")

    contents = file.file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active

    added = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            name = str(row[0]).strip()
            location = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            contact_person = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            phone = str(row[3]).strip() if len(row) > 3 and row[3] else ""

            st = models.Stockist(
                name=name, location=location,
                contact_person=contact_person, phone=phone
            )
            db.add(st)
            added += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()
    return {"message": f"Successfully added {added} stockists", "errors": errors}


@router.post("/products")
def upload_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager"))
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx) are supported")

    contents = file.file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active

    added = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            name = str(row[0]).strip()
            price = float(row[1]) if len(row) > 1 and row[1] else 0.0
            category = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            generic_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            composition = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            dosage = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            packaging = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            manufacturer = str(row[7]).strip() if len(row) > 7 and row[7] else ""
            hsn_code = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            schedule_type = str(row[9]).strip() if len(row) > 9 and row[9] else ""

            # Check if product with this exact name exists to avoid duplicates
            existing = db.query(models.Product).filter(models.Product.name == name).first()
            if existing:
                errors.append(f"Row {i}: Product {name} already exists, skipped")
                continue

            product = models.Product(
                name=name, price=price, category=category,
                generic_name=generic_name, composition=composition,
                dosage=dosage, packaging=packaging,
                manufacturer=manufacturer, hsn_code=hsn_code,
                schedule_type=schedule_type
            )
            db.add(product)
            added += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()
    return {"message": f"Successfully added {added} products", "errors": errors}
