import io
import os
import json
from datetime import datetime, date
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import extract
from sqlalchemy.orm import Session
from database import get_db
from auth import require_role, log_action
from fpdf import FPDF
from openpyxl import Workbook
import models
import schemas

router = APIRouter(prefix="/api/collections", tags=["Collections"])


def clean_pdf_text(text):
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    text = text.replace("—", "-").replace("–", "-").replace("₹", "Rs. ").replace("’", "'").replace("“", '"').replace("”", '"')
    return text.encode("latin-1", "replace").decode("latin-1")


class CollectionsReportPDF(FPDF):
    def header(self):
        self.set_y(10)
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(10, 55, 58) # Deep Forest Teal
        self.cell(0, 8, "SHREYANSH VOLLORA PVT LTD", align="C", new_x="LMARGIN", new_y="NEXT")
        
        self.set_font("Helvetica", "B", 8)
        self.set_text_color(74, 109, 113) # Muted Teal-Gray
        self.cell(0, 4, "Every Step GUIDED BY CARE", align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(2)
        
        self.set_fill_color(10, 55, 58)
        self.rect(10, 26, 190, 1.5, "F")
        self.set_fill_color(20, 168, 156)
        self.rect(10, 27.5, 190, 0.5, "F")
        self.set_y(32)

    def footer(self):
        self.set_y(-15)
        self.set_draw_color(213, 229, 228)
        self.line(10, 282, 200, 282)
        
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(74, 109, 113)
        self.cell(95, 10, "SHREYANSH VOLLORA PVT LTD  Confidential", align="L")
        self.cell(95, 10, f"Page {self.page_no()}/{{nb}}", align="R")

    def section_title(self, title):
        self.ln(4)
        self.set_font("Helvetica", "B", 11)
        self.set_fill_color(10, 55, 58)
        self.set_text_color(255, 255, 255)
        self.cell(0, 8, clean_pdf_text(f"   {title.upper()}"), fill=True, new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0, 0, 0)
        self.ln(2)

    def table_header(self, cols, widths):
        self.set_font("Helvetica", "B", 9)
        self.set_fill_color(227, 239, 239)
        self.set_text_color(10, 55, 58)
        self.set_draw_color(213, 229, 228)
        for col, w in zip(cols, widths):
            self.cell(w, 8, clean_pdf_text(col), border=1, fill=True, align="C")
        self.ln()

    def table_row(self, values, widths, row_idx=0):
        self.set_font("Helvetica", "", 8.5)
        if row_idx % 2 == 1:
            self.set_fill_color(245, 249, 249)
            fill = True
        else:
            fill = False
        self.set_text_color(26, 61, 64)
        self.set_draw_color(225, 236, 235)
        for val, w in zip(values, widths):
            val_str = str(val)[:30] if val is not None else ""
            self.cell(w, 7.5, clean_pdf_text(val_str), border=1, fill=fill, align="C")
        self.ln()


# ─── API Endpoints ───────────────────────────────────────

@router.post("/", response_model=schemas.CollectionOut)
def create_collection(
    col: schemas.CollectionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager"))
):
    db_col = models.Collection(**col.model_dump())
    db.add(db_col)
    db.commit()
    db.refresh(db_col)
    
    # Resolve names for audit logging
    entity_name = "Unknown"
    if db_col.collection_type == "stockist" and db_col.stockist_id:
        st = db.query(models.Stockist).filter(models.Stockist.id == db_col.stockist_id).first()
        entity_name = st.name if st else f"Stockist #{db_col.stockist_id}"
    elif db_col.collection_type == "doctor" and db_col.doctor_id:
        doc = db.query(models.Doctor).filter(models.Doctor.id == db_col.doctor_id).first()
        entity_name = doc.name if doc else f"Doctor #{db_col.doctor_id}"

    log_action(db, current_user.username, "CREATE_COLLECTION", f"Recorded collection of Rs. {db_col.amount:,.2f} from {db_col.collection_type.capitalize()}: {entity_name}")
    
    # Set display names for response
    set_display_names(db_col, db)
    return db_col


@router.get("/", response_model=List[schemas.CollectionOut])
def get_collections(
    month: Optional[int] = Query(None, description="Month of collection (1-12)"),
    year: Optional[int] = Query(None, description="Year of collection (e.g., 2026)"),
    collection_type: Optional[str] = Query(None, description="Filter: 'stockist' or 'doctor'"),
    stockist_id: Optional[int] = Query(None),
    doctor_id: Optional[int] = Query(None),
    employee_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager", "hr"))
):
    query = db.query(models.Collection)
    
    if month is not None:
        query = query.filter(extract('month', models.Collection.date) == month)
    if year is not None:
        query = query.filter(extract('year', models.Collection.date) == year)
    if collection_type:
        query = query.filter(models.Collection.collection_type == collection_type)
    if stockist_id:
        query = query.filter(models.Collection.stockist_id == stockist_id)
    if doctor_id:
        query = query.filter(models.Collection.doctor_id == doctor_id)
    if employee_id:
        query = query.filter(models.Collection.employee_id == employee_id)
        
    cols = query.order_by(models.Collection.date.desc()).all()
    for c in cols:
        set_display_names(c, db)
    return cols


@router.put("/{collection_id}", response_model=schemas.CollectionOut)
def update_collection(
    collection_id: int,
    col_update: schemas.CollectionUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager"))
):
    db_col = db.query(models.Collection).filter(models.Collection.id == collection_id).first()
    if not db_col:
        raise HTTPException(status_code=404, detail="Collection record not found")
        
    for key, value in col_update.model_dump(exclude_unset=True).items():
        setattr(db_col, key, value)
        
    db.commit()
    db.refresh(db_col)
    
    log_action(db, current_user.username, "UPDATE_COLLECTION", f"Updated collection record (ID: {collection_id}) to amount Rs. {db_col.amount:,.2f}")
    set_display_names(db_col, db)
    return db_col


@router.delete("/{collection_id}")
def delete_collection(
    collection_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager"))
):
    db_col = db.query(models.Collection).filter(models.Collection.id == collection_id).first()
    if not db_col:
        raise HTTPException(status_code=404, detail="Collection record not found")
        
    amount = db_col.amount
    db.delete(db_col)
    db.commit()
    
    log_action(db, current_user.username, "DELETE_COLLECTION", f"Deleted collection record ID: {collection_id} (amount Rs. {amount:,.2f})")
    return {"message": "Collection record deleted successfully"}


@router.get("/summary")
def get_collections_summary(
    month: int = Query(..., description="Month of summary (1-12)"),
    year: int = Query(..., description="Year of summary (e.g., 2026)"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager", "hr"))
):
    cols = db.query(models.Collection).filter(
        extract('month', models.Collection.date) == month,
        extract('year', models.Collection.date) == year
    ).all()
    
    total_stockist = 0.0
    total_doctor = 0.0
    
    stockist_breakdown = {}
    doctor_breakdown = {}
    
    for c in cols:
        amount = c.amount or 0.0
        if c.collection_type == "stockist" and c.stockist_id:
            total_stockist += amount
            st = db.query(models.Stockist).filter(models.Stockist.id == c.stockist_id).first()
            name = st.name if st else f"Stockist #{c.stockist_id}"
            stockist_breakdown[name] = stockist_breakdown.get(name, 0.0) + amount
        elif c.collection_type == "doctor" and c.doctor_id:
            total_doctor += amount
            doc = db.query(models.Doctor).filter(models.Doctor.id == c.doctor_id).first()
            name = doc.name if doc else f"Doctor #{c.doctor_id}"
            doctor_breakdown[name] = doctor_breakdown.get(name, 0.0) + amount
            
    return {
        "month": month,
        "year": year,
        "total_stockist": total_stockist,
        "total_doctor": total_doctor,
        "grand_total": total_stockist + total_doctor,
        "stockist_breakdown": [{"name": k, "amount": v} for k, v in stockist_breakdown.items()],
        "doctor_breakdown": [{"name": k, "amount": v} for k, v in doctor_breakdown.items()]
    }


@router.get("/export/excel")
def export_collections_excel(
    month: int = Query(...),
    year: int = Query(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager", "hr"))
):
    cols = db.query(models.Collection).filter(
        extract('month', models.Collection.date) == month,
        extract('year', models.Collection.date) == year
    ).order_by(models.Collection.date.desc()).all()
    
    for c in cols:
        set_display_names(c, db)
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Collections Report"
    
    # Styled Header
    ws.append(["SHREYANSH VOLLORA PVT LTD"])
    ws.append([f"MONTHLY COLLECTIONS REPORT - {datetime(year, month, 1).strftime('%B %Y').upper()}"])
    ws.append([])
    
    ws.append(["Date", "Type", "Party Name", "Collected By (MR)", "Payment Mode", "Amount (Rs.)", "Remarks"])
    
    grand_total = 0.0
    for c in cols:
        party_name = c.stockist_name if c.collection_type == "stockist" else c.doctor_name
        amount = c.amount or 0.0
        grand_total += amount
        ws.append([
            c.date.strftime("%Y-%m-%d") if c.date else "",
            c.collection_type.capitalize(),
            party_name or "Unknown",
            c.employee_name or "-",
            c.payment_mode,
            amount,
            c.remarks or ""
        ])
        
    ws.append([])
    ws.append(["GRAND TOTAL COLLECTIONS", "", "", "", "", grand_total])
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    month_name = datetime(year, month, 1).strftime("%B_%Y")
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=collections_report_{month_name}.xlsx"}
    )


@router.get("/export/pdf")
def export_collections_pdf(
    month: int = Query(...),
    year: int = Query(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin", "manager", "hr"))
):
    cols = db.query(models.Collection).filter(
        extract('month', models.Collection.date) == month,
        extract('year', models.Collection.date) == year
    ).order_by(models.Collection.date.desc()).all()
    
    for c in cols:
        set_display_names(c, db)
        
    pdf = CollectionsReportPDF()
    pdf.add_page()
    
    # Title Block
    pdf.set_y(35)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(10, 55, 58)
    month_str = datetime(year, month, 1).strftime("%B %Y")
    pdf.cell(0, 8, clean_pdf_text(f"MONTHLY COLLECTIONS REPORT: {month_str.upper()}"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    
    # Summary calculations
    total_st = sum(c.amount for c in cols if c.collection_type == "stockist")
    total_doc = sum(c.amount for c in cols if c.collection_type == "doctor")
    grand_total = total_st + total_doc
    
    # Summary Grid Info
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(26, 61, 64)
    pdf.cell(60, 6, clean_pdf_text(f"Total Stockist Collections: Rs. {total_st:,.2f}"), border=0)
    pdf.cell(60, 6, clean_pdf_text(f"Total Doctor Collections: Rs. {total_doc:,.2f}"), border=0)
    pdf.cell(70, 6, clean_pdf_text(f"Grand Total Collected: Rs. {grand_total:,.2f}"), border=0)
    pdf.ln(8)
    
    # Detailed Collection Table
    pdf.section_title("Detailed Payment Collections List")
    
    headers = ["Date", "Type", "Party Name", "Collected By (MR)", "Mode", "Amount (Rs.)"]
    widths = [24, 18, 55, 48, 20, 25]
    pdf.table_header(headers, widths)
    
    for idx, c in enumerate(cols):
        party_name = c.stockist_name if c.collection_type == "stockist" else c.doctor_name
        values = [
            c.date.strftime("%Y-%m-%d") if c.date else "",
            c.collection_type.capitalize(),
            party_name or "-",
            c.employee_name or "-",
            c.payment_mode,
            f"{c.amount:,.2f}"
        ]
        pdf.table_row(values, widths, idx)
        
    # PDF Stream Response
    pdf_bytes = pdf.output()
    file_stream = io.BytesIO(pdf_bytes)
    
    month_name = datetime(year, month, 1).strftime("%B_%Y")
    return StreamingResponse(
        file_stream,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=collections_report_{month_name}.pdf"}
    )


# ─── Helper Functions ────────────────────────────────────

def set_display_names(col: models.Collection, db: Session):
    col.stockist_name = "-"
    col.doctor_name = "-"
    col.employee_name = "-"
    
    if col.collection_type == "stockist" and col.stockist_id:
        st = db.query(models.Stockist).filter(models.Stockist.id == col.stockist_id).first()
        if st:
            col.stockist_name = st.name
            
    elif col.collection_type == "doctor" and col.doctor_id:
        doc = db.query(models.Doctor).filter(models.Doctor.id == col.doctor_id).first()
        if doc:
            col.doctor_name = doc.name
            
    if col.employee_id:
        emp = db.query(models.Employee).filter(models.Employee.id == col.employee_id).first()
        if emp:
            col.employee_name = emp.name
